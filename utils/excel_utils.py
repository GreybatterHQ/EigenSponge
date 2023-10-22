import os
from openpyxl import load_workbook, Workbook
import pandas as pd

def save_dataFrame_to_excel(excel_file, sheet_name, df, directory=None):
    if directory:
        if not os.path.exists(directory):
            os.makedirs(directory)
        excel_file = os.path.join(directory, excel_file)

    # check if the file exists
    try:
        print('loading excel file')
        workbook = load_workbook(excel_file)
        writer = pd.ExcelWriter(excel_file, engine='openpyxl')
        writer.book = workbook
        if sheet_name in workbook.sheetnames:
            # append the data to the sheet
            sheet = workbook[sheet_name]
            writer.sheets = {ws.title: ws for ws in workbook.worksheets}
            start_row = sheet.max_row
            df.to_excel(writer, sheet_name, startrow=start_row, header=None, index=False)
        else:
            # create a new sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    except FileNotFoundError:
        print('creating new file', excel_file)
        # create a new workbook
        writer = pd.ExcelWriter(excel_file, engine='openpyxl')
        writer.book = Workbook()
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    writer.save()
    writer.close()
    print(f'saved data to {excel_file} under {sheet_name} sheet')