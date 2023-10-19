import os
import json
import requests
import pandas as pd
from openpyxl import load_workbook, Workbook
import numpy as np
from utils.http_utils import request_handler

def amazon_search(api_basic_auth_token, query, url, headers):
    payload = {
      "target": "amazon_search",
      "query": str(query),
      "parse": True,
      "domain": "in",
      "device_type": "desktop",
      "page_from": "1"
      }

    response_json = request_handler(url, 'POST', data=payload, headers=headers)
    results = response_json.get('results')[0].get('content').get('results')

    result_directory = 'output'
    fileName = 'amazon_search.xlsx'

    df_paid = pd.DataFrame(results.get('paid'))
    df_organic= pd.DataFrame(results.get('organic'))
    df_amazon_choices= pd.DataFrame(results.get('amazons_choices'))

    save_dataFrame_to_excel(fileName, 'paid', df_paid, result_directory)
    save_dataFrame_to_excel(fileName, 'organic', df_organic, result_directory)
    save_dataFrame_to_excel(fileName, 'amazon_choices', df_amazon_choices, result_directory)
    

def save_dataFrame_to_excel(excel_file, sheet_name, df, directory=None):
    if directory:
        if not os.path.exists(directory):
            os.makedirs(directory)
        excel_file = os.path.join(directory, excel_file)

    # check if the file exists
    try:
        workbook = load_workbook(excel_file)
        writer = pd.ExcelWriter(excel_file, engine='openpyxl')
        writer.book = workbook
    except FileNotFoundError:
        # create a new workbook
        writer = pd.ExcelWriter(excel_file, engine='openpyxl')
        writer.book = Workbook()
        

    df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.save()
    writer.close()

def amazon_product(api_basic_auth_token, query, url, headers):
    payload = {
      "target": "amazon_product",
      "query": str(query),
      "domain": "in",
      "device_type": "desktop",
      "parse": True
      }
    # response = requests.post(url, json=payload, headers=headers)
    response_json = request_handler(url, 'POST', payload, headers)
    # print(response.text)
    result = response_json.get('results')[0].get('content')
    df = pd.DataFrame(result)
    df.to_excel('output_excel_amazon_product'+str(query)+'.xlsx', index=False)

def amazon_pricing(api_basic_auth_token, query, url, headers):
    payload = {
      "target": "amazon_pricing",
      "query": str(query),
      "domain": "in",
      "device_type": "desktop",
      "parse": True
      }
    # response = requests.post(url, json=payload, headers=headers)
    response_json = request_handler(url, 'POST', payload, headers)
    # print(response.text)
    result = response_json.get('results')[0].get('content')
    df = pd.DataFrame(result)
    df.to_excel('output_excel_amazon_pricing'+str(query)+'.xlsx')

def amazon_questions(api_basic_auth_token, query, url, headers):
    payload = {
      "target": "amazon_questions",
      "query": str(query),
      "domain": "in",
      "device_type": "desktop",
      "parse": True
      }
    # response = requests.post(url, json=payload, headers=headers)
    response_json = request_handler(url, 'POST', payload, headers)
    # print(response.text)
    questions = response.json().get('results')[0].get('content').get('questions')
    df = pd.DataFrame(questions)
    df.to_excel('output_excel_amazon_questions'+str(query)+'.xlsx', index=False)

def amazon_reviews(api_basic_auth_token, query, url, headers):
    payload = {
      "target": "amazon_reviews",
      "query": str(query),
      "domain": "in",
      "device_type": "desktop",
      "parse": True
      }
    # response = requests.post(url, json=payload, headers=headers)
    response_json = request_handler(url, 'POST', payload, headers)
    # print(response.text)
    reviews = response_json.get('results')[0].get('content').get('reviews')
    df = pd.DataFrame(reviews)
    df.to_excel('output_excel_amazon_reviews'+str(query)+'.xlsx', index=False)


def init_api_auth(config_path):
  with open(config_path,"r") as config_file:
    config = json.load(config_file)
  return config['api_basic_auth_token']


def main():
  api_basic_auth = init_api_auth("config.json")
  url = "https://scrape.smartproxy.com/v1/tasks"
  headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "authorization": "Basic "+ api_basic_auth
    }

  excel_amazon_search= pd.read_excel('input_amazon_search.xlsx')
  excel_amazon_product= pd.read_excel('input_amazon_product.xlsx')

  print('printing the search input:', excel_amazon_search)
  print('amazon product:',excel_amazon_product)

  #for row in excel_amazon_search.itertuples():
  #   print("\n\n\n")
  for row in excel_amazon_product.itertuples():    
    amazon_search(api_basic_auth_token=api_basic_auth,query=row[1], url=url, headers=headers)
    # print("\n\n\n")
    #amazon_product(api_basic_auth_token=api_basic_auth,query=row[1], url=url, headers=headers)
    # print("\n\n\n")
    # amazon_pricing(api_basic_auth_token=api_basic_auth,query=row[1], url=url, headers=headers)
    # print("\n\n\n")
    # amazon_questions(api_basic_auth_token=api_basic_auth,query=row[1], url=url, headers=headers)
    # print("\n\n\n")
    # amazon_reviews(api_basic_auth_token=api_basic_auth,query=row[1], url=url, headers=headers)
    exit()


main()
    
